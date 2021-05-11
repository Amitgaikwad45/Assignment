#read data from excel
import openpyxl
path="F:/PositionSalary.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active #workbook.get_sheet_by_name("PositionSalary")
row=sheet.max_row
col=sheet.max_column

print(row)
print(col)

for r in range(1,row+1):
    for c in range(1,col+1):
        print(sheet.cell(row=r,column=c).value,end="   ")
    print()